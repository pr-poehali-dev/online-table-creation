'''
Business: Export spreadsheet to CSV or Excel format
Args: event with httpMethod, queryStringParameters (sheet_id, format); context with request_id
Returns: HTTP response with file download or error
'''

import json
import os
import io
import csv
import base64
from typing import Dict, Any
import psycopg2
from psycopg2.extras import RealDictCursor
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

def get_db_connection():
    """Create database connection using simple query protocol"""
    dsn = os.environ.get('DATABASE_URL')
    return psycopg2.connect(dsn)

def handler(event: Dict[str, Any], context: Any) -> Dict[str, Any]:
    method: str = event.get('httpMethod', 'GET')
    
    if method == 'OPTIONS':
        return {
            'statusCode': 200,
            'headers': {
                'Access-Control-Allow-Origin': '*',
                'Access-Control-Allow-Methods': 'GET, OPTIONS',
                'Access-Control-Allow-Headers': 'Content-Type',
                'Access-Control-Max-Age': '86400'
            },
            'body': ''
        }
    
    if method != 'GET':
        return {
            'statusCode': 405,
            'headers': {'Access-Control-Allow-Origin': '*'},
            'body': json.dumps({'error': 'Method not allowed'})
        }
    
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    
    try:
        params = event.get('queryStringParameters') or {}
        sheet_id = params.get('sheet_id', '1')
        export_format = params.get('format', 'csv').lower()
        
        cur.execute(f"SELECT row_index, col_index, value FROM cells WHERE sheet_id = {sheet_id} ORDER BY row_index, col_index")
        cells = cur.fetchall()
        
        if not cells:
            return {
                'statusCode': 404,
                'headers': {
                    'Content-Type': 'application/json',
                    'Access-Control-Allow-Origin': '*'
                },
                'body': json.dumps({'error': 'No data found'})
            }
        
        max_row = max(cell['row_index'] for cell in cells)
        max_col = max(cell['col_index'] for cell in cells)
        
        grid = [['' for _ in range(max_col + 1)] for _ in range(max_row + 1)]
        for cell in cells:
            grid[cell['row_index']][cell['col_index']] = cell['value'] or ''
        
        if export_format == 'csv':
            output = io.StringIO()
            writer = csv.writer(output)
            writer.writerows(grid)
            csv_content = output.getvalue()
            
            return {
                'statusCode': 200,
                'headers': {
                    'Content-Type': 'text/csv',
                    'Content-Disposition': 'attachment; filename="table_export.csv"',
                    'Access-Control-Allow-Origin': '*'
                },
                'isBase64Encoded': False,
                'body': csv_content
            }
        
        elif export_format == 'excel':
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Exported Data"
            
            header_fill = PatternFill(start_color='8B5CF6', end_color='8B5CF6', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF')
            
            for row_idx, row_data in enumerate(grid, start=1):
                for col_idx, value in enumerate(row_data, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    if row_idx == 1:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal='center')
            
            cols_list = list(ws.columns)
            for col in cols_list:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                ws.column_dimensions[column].width = min(max_length + 2, 50)
            
            output = io.BytesIO()
            wb.save(output)
            excel_content = output.getvalue()
            excel_b64 = base64.b64encode(excel_content).decode('utf-8')
            
            return {
                'statusCode': 200,
                'headers': {
                    'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    'Content-Disposition': 'attachment; filename="table_export.xlsx"',
                    'Access-Control-Allow-Origin': '*'
                },
                'isBase64Encoded': True,
                'body': excel_b64
            }
        
        else:
            return {
                'statusCode': 400,
                'headers': {
                    'Content-Type': 'application/json',
                    'Access-Control-Allow-Origin': '*'
                },
                'body': json.dumps({'error': 'Invalid format. Use csv or excel'})
            }
    
    finally:
        cur.close()
        conn.close()